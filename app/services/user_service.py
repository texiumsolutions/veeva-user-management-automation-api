"""User service for business logic"""
from models.user import User
from pymongo.errors import DuplicateKeyError
from typing import List, Dict
import openpyxl
from io import BytesIO


class UserService:
    """Service for user operations"""
    
    @staticmethod
    def create_user_from_json(user_data: dict) -> Dict:
        """
        Create a user from JSON ingest data
        Extracts email and name from nested structure
        """
        try:
            email = user_data.get("user_email__v")
            first_name = user_data.get("user_first_name__v")
            last_name = user_data.get("user_last_name__v")
            user_name = user_data.get("user_name__v")
            
            user_info = {
                "email": email,
                "first_name": first_name,
                "last_name": last_name,
                "user_name": user_name or email
            }
            
            user_id = User.insert_user(user_info)
            return {
                "success": True,
                "message": f"User {email} created successfully",
                "user_id": str(user_id)
            }
        except DuplicateKeyError:
            return {
                "success": False,
                "message": f"User with email {email} already exists",
                "user_id": None
            }
        except Exception as e:
            return {
                "success": False,
                "message": str(e),
                "user_id": None
            }
    
    @staticmethod
    def create_users_from_json_payload(payload: dict) -> Dict:
        """
        Create multiple users from JSON payload
        """
        users = payload.get("users", [])
        results = {
            "total": len(users),
            "successful": 0,
            "failed": 0,
            "details": []
        }
        
        for user_data in users:
            result = UserService.create_user_from_json(user_data)
            if result["success"]:
                results["successful"] += 1
            else:
                results["failed"] += 1
            results["details"].append(result)
        
        return results
    
    @staticmethod
    def create_users_from_excel(file_content: bytes) -> Dict:
        """
        Create users from Excel file
        Expects columns: user_name__v, user_first_name__v, user_last_name__v, user_email__v
        """
        try:
            # Load Excel file
            excel_file = BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active
            
            results = {
                "total": 0,
                "successful": 0,
                "failed": 0,
                "details": []
            }
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(cell.value)
            
            # Map column indices
            col_map = {header: idx for idx, header in enumerate(headers, 1)}
            
            # Process data rows (skip header)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # Get values from cells
                    user_name = worksheet.cell(row_idx, col_map.get("user_name__v", 1)).value
                    first_name = worksheet.cell(row_idx, col_map.get("user_first_name__v", 2)).value
                    last_name = worksheet.cell(row_idx, col_map.get("user_last_name__v", 3)).value
                    email = worksheet.cell(row_idx, col_map.get("user_email__v", 4)).value
                    
                    # Skip empty rows
                    if not email:
                        continue
                    
                    results["total"] += 1
                    
                    user_data = {
                        "user_name__v": user_name or email,
                        "user_first_name__v": first_name or "",
                        "user_last_name__v": last_name or "",
                        "user_email__v": email
                    }
                    
                    result = UserService.create_user_from_json(user_data)
                    if result["success"]:
                        results["successful"] += 1
                    else:
                        results["failed"] += 1
                    results["details"].append({
                        "row": row_idx,
                        **result
                    })
                
                except Exception as e:
                    results["failed"] += 1
                    results["details"].append({
                        "row": row_idx,
                        "success": False,
                        "message": str(e)
                    })
            
            return results
        
        except Exception as e:
            return {
                "total": 0,
                "successful": 0,
                "failed": 0,
                "details": [],
                "error": str(e)
            }
    
    @staticmethod
    def get_all_users() -> List[Dict]:
        """Get all users from database"""
        return User.find_all_users()
